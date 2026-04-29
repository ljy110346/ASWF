from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

try:
    import torch
    from torch.utils.data import Dataset
except ModuleNotFoundError:  # pragma: no cover - exercised only when torch is unavailable.
    torch = None

    class Dataset:  # type: ignore[no-redef]
        """Fallback stub so config and discovery utilities can still import this module."""


FILE_KEYS = ("ir_normal", "raman_normal", "ir_abnormal", "raman_abnormal")


@dataclass(frozen=True)
class WaveNumberAxis:
    modality: str
    values: Optional[np.ndarray]
    path: Optional[Path]
    sheet_name: Optional[str]
    is_valid: bool
    warning: Optional[str] = None


@dataclass(frozen=True)
class DiseaseFiles:
    disease: str
    root: Path
    spectral_files: Dict[str, Path]
    axis_files: Dict[str, Optional[Path]]
    chosen_sheets: Dict[str, Optional[str]]
    chosen_axis_sheets: Dict[str, Optional[str]]


@dataclass
class DiseaseDataBundle:
    disease: str
    ir: np.ndarray
    raman: np.ndarray
    labels: np.ndarray
    sample_ids: List[str]
    files: DiseaseFiles
    wave_numbers: Dict[str, WaveNumberAxis]
    warnings: List[str] = field(default_factory=list)

    @property
    def num_samples(self) -> int:
        return int(self.labels.shape[0])

    @property
    def class_distribution(self) -> Dict[int, int]:
        values, counts = np.unique(self.labels, return_counts=True)
        return {int(value): int(count) for value, count in zip(values, counts)}

    @property
    def modality_dims(self) -> Dict[str, int]:
        return {"ir": int(self.ir.shape[1]), "raman": int(self.raman.shape[1])}

    def summary(self) -> Dict[str, object]:
        return {
            "disease": self.disease,
            "num_samples": self.num_samples,
            "class_distribution": self.class_distribution,
            "modality_dims": self.modality_dims,
            "warnings": list(self.warnings),
            "chosen_sheet_ir": self.files.chosen_sheets.get("ir"),
            "chosen_sheet_raman": self.files.chosen_sheets.get("raman"),
        }


class PairedSpectraDataset(Dataset):
    def __init__(
        self,
        ir: np.ndarray,
        raman: np.ndarray,
        labels: np.ndarray,
        indices: Optional[Iterable[int]] = None,
    ) -> None:
        if torch is None:
            raise ModuleNotFoundError("PyTorch is required to construct PairedSpectraDataset.")
        if indices is None:
            indices = np.arange(labels.shape[0])

        self.indices = np.asarray(list(indices), dtype=np.int64)
        self.ir = torch.as_tensor(ir[self.indices], dtype=torch.float32)
        self.raman = torch.as_tensor(raman[self.indices], dtype=torch.float32)
        self.labels = torch.as_tensor(labels[self.indices], dtype=torch.long)

    def __len__(self) -> int:
        return int(self.indices.shape[0])

    def __getitem__(self, index: int) -> Dict[str, torch.Tensor]:
        return {
            "ir": self.ir[index],
            "raman": self.raman[index],
            "label": self.labels[index],
        }


def _pick_sheet(path: Path, preferred_sheet: Optional[str]) -> Optional[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if preferred_sheet is not None:
        if preferred_sheet not in workbook.sheetnames:
            raise ValueError(f"Requested sheet '{preferred_sheet}' is missing in {path}. Available: {workbook.sheetnames}")
        return preferred_sheet
    return workbook.sheetnames[0] if workbook.sheetnames else None


def _read_excel_matrix(path: Path, sheet_name: Optional[str]) -> np.ndarray:
    frame = pd.read_excel(path, header=None, sheet_name=sheet_name)
    return frame.to_numpy(dtype=np.float32)


def _classify_label(stem: str, keywords: Dict[str, List[str]]) -> Optional[str]:
    if any(keyword in stem for keyword in keywords["abnormal"]):
        return "abnormal"
    if any(keyword in stem for keyword in keywords["normal"]):
        return "normal"
    return None


def _classify_modality(stem: str, keywords: Dict[str, List[str]]) -> Optional[str]:
    if any(keyword in stem for keyword in keywords["raman"]):
        return "raman"
    if any(keyword in stem for keyword in keywords["ir"] if keyword != "ir"):
        return "ir"
    if stem.startswith("ir") or "_ir" in stem or "-ir" in stem or "ir_" in stem or "ir-" in stem:
        return "ir"
    return None


def _resolve_explicit_file(root: Path, file_name: Optional[str], disease: str, key: str) -> Optional[Path]:
    if file_name is None:
        return None
    path = root / file_name
    if not path.is_file():
        raise FileNotFoundError(f"Configured file for disease='{disease}', key='{key}' does not exist: {path}")
    return path


def discover_disease_files(
    data_root: Path,
    disease: str,
    data_config,
) -> DiseaseFiles:
    disease_root = data_root / disease
    if not disease_root.is_dir():
        raise FileNotFoundError(f"Disease directory does not exist: {disease_root}")

    explicit_mapping = data_config.file_mappings.get(disease)
    sheet_overrides = data_config.sheet_overrides.get(disease)
    keywords = data_config.discovery_keywords

    spectral_files: Dict[str, Path] = {}
    for key in FILE_KEYS:
        explicit = _resolve_explicit_file(
            disease_root,
            getattr(explicit_mapping, key) if explicit_mapping is not None else None,
            disease,
            key,
        )
        if explicit is not None:
            spectral_files[key] = explicit
            continue

        matches: List[Path] = []
        for candidate in sorted(disease_root.glob("*.xlsx")):
            if candidate.name.lower().startswith("wavenumber-"):
                continue
            stem = candidate.stem.lower()
            class_name = _classify_label(stem, keywords)
            modality = _classify_modality(stem, keywords)
            if class_name is None or modality is None:
                continue
            if key == f"{modality}_{class_name}":
                matches.append(candidate)

        if len(matches) != 1:
            raise ValueError(
                f"Auto discovery for disease='{disease}', key='{key}' expected exactly one match, "
                f"found {len(matches)}: {[path.name for path in matches]}"
            )
        spectral_files[key] = matches[0]

    axis_files: Dict[str, Optional[Path]] = {}
    for modality, explicit_key, default_name in (
        ("ir", "wavenumber_ir", "Wavenumber-ir.xlsx"),
        ("raman", "wavenumber_raman", "Wavenumber-raman.xlsx"),
    ):
        explicit_name = getattr(explicit_mapping, explicit_key) if explicit_mapping is not None else None
        if explicit_name is not None:
            axis_files[modality] = _resolve_explicit_file(disease_root, explicit_name, disease, explicit_key)
            continue
        default_path = disease_root / default_name
        axis_files[modality] = default_path if default_path.is_file() else None

    chosen_sheets = {
        "ir": _pick_sheet(spectral_files["ir_normal"], getattr(sheet_overrides, "ir", None) if sheet_overrides else None),
        "raman": _pick_sheet(
            spectral_files["raman_normal"],
            getattr(sheet_overrides, "raman", None) if sheet_overrides else None,
        ),
    }
    chosen_axis_sheets = {
        "ir": _pick_sheet(axis_files["ir"], getattr(sheet_overrides, "wavenumber_ir", None) if sheet_overrides and axis_files["ir"] else None)
        if axis_files["ir"] is not None
        else None,
        "raman": _pick_sheet(
            axis_files["raman"],
            getattr(sheet_overrides, "wavenumber_raman", None) if sheet_overrides and axis_files["raman"] else None,
        )
        if axis_files["raman"] is not None
        else None,
    }

    return DiseaseFiles(
        disease=disease,
        root=disease_root,
        spectral_files=spectral_files,
        axis_files=axis_files,
        chosen_sheets=chosen_sheets,
        chosen_axis_sheets=chosen_axis_sheets,
    )


def _validate_sheet_consistency(path: Path, expected_sheet: Optional[str]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if expected_sheet is not None and expected_sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet '{expected_sheet}' is missing in {path}. Available: {workbook.sheetnames}")


def _build_axis(
    modality: str,
    axis_path: Optional[Path],
    sheet_name: Optional[str],
    expected_dim: int,
    mismatch_policy: str,
) -> WaveNumberAxis:
    if axis_path is None:
        return WaveNumberAxis(
            modality=modality,
            values=None,
            path=None,
            sheet_name=None,
            is_valid=False,
            warning=f"Wave number axis is missing for modality '{modality}'.",
        )

    values = _read_excel_matrix(axis_path, sheet_name).reshape(-1)
    if values.size != expected_dim:
        warning = (
            f"Wave number axis length mismatch for modality '{modality}': "
            f"expected {expected_dim}, got {values.size} ({axis_path.name}, sheet={sheet_name})."
        )
        if mismatch_policy == "error":
            raise ValueError(warning)
        return WaveNumberAxis(
            modality=modality,
            values=values.astype(np.float32, copy=False),
            path=axis_path,
            sheet_name=sheet_name,
            is_valid=False,
            warning=warning,
        )

    return WaveNumberAxis(
        modality=modality,
        values=values.astype(np.float32, copy=False),
        path=axis_path,
        sheet_name=sheet_name,
        is_valid=True,
        warning=None,
    )


def load_disease_bundle(data_root: Path, disease: str, config) -> DiseaseDataBundle:
    files = discover_disease_files(data_root=data_root, disease=disease, data_config=config.data)

    for key, path in files.spectral_files.items():
        modality = "raman" if key.startswith("raman") else "ir"
        _validate_sheet_consistency(path, files.chosen_sheets[modality])

    ir_normal = _read_excel_matrix(files.spectral_files["ir_normal"], files.chosen_sheets["ir"])
    raman_normal = _read_excel_matrix(files.spectral_files["raman_normal"], files.chosen_sheets["raman"])
    ir_abnormal = _read_excel_matrix(files.spectral_files["ir_abnormal"], files.chosen_sheets["ir"])
    raman_abnormal = _read_excel_matrix(files.spectral_files["raman_abnormal"], files.chosen_sheets["raman"])

    if ir_normal.shape[0] != raman_normal.shape[0]:
        raise ValueError(
            f"Normal pair count mismatch for disease='{disease}': ir={ir_normal.shape[0]}, raman={raman_normal.shape[0]}"
        )
    if ir_abnormal.shape[0] != raman_abnormal.shape[0]:
        raise ValueError(
            f"Abnormal pair count mismatch for disease='{disease}': "
            f"ir={ir_abnormal.shape[0]}, raman={raman_abnormal.shape[0]}"
        )

    ir = np.vstack([ir_normal, ir_abnormal]).astype(np.float32, copy=False)
    raman = np.vstack([raman_normal, raman_abnormal]).astype(np.float32, copy=False)
    labels = np.concatenate(
        [
            np.zeros(ir_normal.shape[0], dtype=np.int64),
            np.ones(ir_abnormal.shape[0], dtype=np.int64),
        ]
    )
    sample_ids = [f"normal_{index:04d}" for index in range(ir_normal.shape[0])] + [
        f"abnormal_{index:04d}" for index in range(ir_abnormal.shape[0])
    ]

    wave_numbers = {
        "ir": _build_axis(
            modality="ir",
            axis_path=files.axis_files["ir"],
            sheet_name=files.chosen_axis_sheets["ir"],
            expected_dim=ir.shape[1],
            mismatch_policy=config.data.axis_mismatch_policy,
        ),
        "raman": _build_axis(
            modality="raman",
            axis_path=files.axis_files["raman"],
            sheet_name=files.chosen_axis_sheets["raman"],
            expected_dim=raman.shape[1],
            mismatch_policy=config.data.axis_mismatch_policy,
        ),
    }

    warnings = [axis.warning for axis in wave_numbers.values() if axis.warning]
    return DiseaseDataBundle(
        disease=disease,
        ir=ir,
        raman=raman,
        labels=labels,
        sample_ids=sample_ids,
        files=files,
        wave_numbers=wave_numbers,
        warnings=warnings,
    )
